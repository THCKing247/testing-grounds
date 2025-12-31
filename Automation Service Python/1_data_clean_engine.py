"""
Apex Data Clean Engine

Cleans, standardizes, and fixes messy data files from various sources.
Supports CSV, Excel, JSON, TSV, and CRM exports (Salesforce, HubSpot, etc.).

Offline behavior:
- CSV/TSV supported via stdlib `csv`
- XLSX/XLS supported if `openpyxl` is installed (optional)
- JSON supported via stdlib `json`
"""

from __future__ import annotations

import csv
import dataclasses
import datetime as _dt
import io
import json
import os
import re
import typing as t

from shared_utils import ServiceError, slugify_header, utc_now_iso


@dataclasses.dataclass
class DataCleanReport:
    rows_in: int
    rows_out: int
    columns_in: int
    columns_out: int
    header_map: dict[str, str]
    fixes: dict[str, int]
    started_at: str
    finished_at: str
    file_type: str = "csv"
    crm_detected: str | None = None
    field_mappings: dict[str, str] = dataclasses.field(default_factory=dict)
    duplicates_removed: int = 0
    irrelevant_rows_removed: int = 0


class ApexDataCleanEngine:
    """
    Cleans, standardizes, and fixes messy data files from various sources.
    Supports CSV, Excel, JSON, TSV, and CRM exports (Salesforce, HubSpot, etc.).

    Offline behavior:
    - CSV/TSV supported via stdlib `csv`
    - XLSX/XLS supported if `openpyxl` is installed (optional)
    - JSON supported via stdlib `json`
    """

    # CRM field mappings - maps common CRM field names to standardized names
    CRM_FIELD_MAPPINGS = {
        "salesforce": {
            "firstname": "first_name",
            "lastname": "last_name",
            "email": "email",
            "phone": "phone",
            "mobilephone": "mobile_phone",
            "mailingstreet": "address_street",
            "mailingcity": "address_city",
            "mailingstate": "address_state",
            "mailingpostalcode": "address_postal_code",
            "mailingcountry": "address_country",
            "company": "company_name",
            "title": "job_title",
            "lead_source": "lead_source",
            "status": "status",
            "createddate": "created_date",
            "lastmodifieddate": "last_modified_date",
        },
        "hubspot": {
            "firstname": "first_name",
            "lastname": "last_name",
            "email": "email",
            "phone": "phone",
            "mobilephone": "mobile_phone",
            "address": "address_street",
            "city": "address_city",
            "state": "address_state",
            "zip": "address_postal_code",
            "country": "address_country",
            "company": "company_name",
            "jobtitle": "job_title",
            "leadsource": "lead_source",
            "lifecyclestage": "status",
            "createdate": "created_date",
            "hs_lastmodifieddate": "last_modified_date",
        },
        "pipedrive": {
            "first_name": "first_name",
            "last_name": "last_name",
            "email": "email",
            "phone": "phone",
            "mobile": "mobile_phone",
            "org_name": "company_name",
            "owner_name": "owner_name",
            "add_time": "created_date",
            "update_time": "last_modified_date",
        },
    }

    _date_patterns = (
        # 2025-12-29
        re.compile(r"^\d{4}-\d{2}-\d{2}$"),
        # 12/29/2025, 12-29-2025
        re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$"),
        # Salesforce format: 2025-12-29T10:30:00.000Z
        re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}"),
    )

    def clean_csv_text(
        self,
        csv_text: str,
        *,
        delimiter: str = ",",
        normalize_headers: bool = True,
        drop_empty_rows: bool = True,
    ) -> tuple[str, DataCleanReport]:
        started = utc_now_iso()
        fixes: dict[str, int] = {
            "trimmed_cells": 0,
            "normalized_headers": 0,
            "normalized_dates": 0,
            "normalized_numbers": 0,
            "empties_to_blank": 0,
            "dropped_empty_rows": 0,
            "duplicates_removed": 0,
            "irrelevant_rows_removed": 0,
        }

        inp = io.StringIO(csv_text)
        reader = csv.reader(inp, delimiter=delimiter)
        rows = list(reader)
        if not rows:
            raise ServiceError("CSV appears to be empty.")

        raw_headers = rows[0]
        data_rows = rows[1:]

        header_map: dict[str, str] = {}
        headers_out: list[str] = []
        for h in raw_headers:
            h2 = slugify_header(h) if normalize_headers else h.strip()
            header_map[h] = h2
            headers_out.append(h2)
            if h2 != h:
                fixes["normalized_headers"] += 1

        def norm_cell(val: str) -> str:
            if val is None:
                fixes["empties_to_blank"] += 1
                return ""
            v = str(val)
            v2 = v.strip()
            if v2 != v:
                fixes["trimmed_cells"] += 1
            v = v2
            if not v:
                return ""

            # numbers like "1,234.50" -> "1234.50"
            if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", v):
                fixes["normalized_numbers"] += 1
                return v.replace(",", "")

            # dates
            for pat in self._date_patterns:
                if pat.match(v):
                    iso = self._try_parse_date_to_iso(v)
                    if iso and iso != v:
                        fixes["normalized_dates"] += 1
                        return iso
                    return v
            return v

        def reconcile_row_length(row: list[str], target_cols: int) -> list[str]:
            """
            Best-effort fix for malformed CSV rows where delimiters appear inside values
            but the value wasn't quoted (e.g., 1,234.50).
            """
            if len(row) == target_cols:
                return row
            if len(row) < target_cols:
                return (row + [""] * target_cols)[:target_cols]

            # Attempt to merge numeric thousands separators: ["1","234.50"] -> ["1,234.50"]
            rr = list(row)
            numeric_left = re.compile(r"^-?\d{1,3}$")
            numeric_mid = re.compile(r"^\d{3}$")
            numeric_right = re.compile(r"^\d{3}(\.\d+)?$")

            i = 0
            while len(rr) > target_cols and i < len(rr) - 1:
                a, b = rr[i].strip(), rr[i + 1].strip()
                if numeric_left.match(a) and (numeric_mid.match(b) or numeric_right.match(b)):
                    rr[i : i + 2] = [f"{a},{b}"]
                    fixes["normalized_numbers"] += 1
                    continue
                i += 1

            # If still too long, merge the tail into the last column.
            if len(rr) > target_cols:
                head = rr[: target_cols - 1]
                tail = rr[target_cols - 1 :]
                rr = head + [delimiter.join(tail)]
            return rr[:target_cols]

        # Process and clean rows
        cleaned_rows: list[list[str]] = []
        for r in data_rows:
            rr = reconcile_row_length(list(r), len(raw_headers))
            rr2 = [norm_cell(c) for c in rr]
            
            # Drop completely empty rows
            if drop_empty_rows and all(c == "" for c in rr2):
                fixes["dropped_empty_rows"] += 1
                continue
            
            # Filter irrelevant rows (rows with too many empty cells or test data)
            if self._is_irrelevant_row(rr2, len(headers_out)):
                fixes["irrelevant_rows_removed"] += 1
                continue
            
            cleaned_rows.append(rr2)
        
        # Remove duplicates
        seen_rows: set[tuple[str, ...]] = set()
        deduplicated_rows: list[list[str]] = []
        for row in cleaned_rows:
            # Create a tuple for hashing (normalize whitespace)
            row_tuple = tuple(c.strip().lower() if c else "" for c in row)
            if row_tuple not in seen_rows:
                seen_rows.add(row_tuple)
                deduplicated_rows.append(row)
            else:
                fixes["duplicates_removed"] += 1
        
        out_rows: list[list[str]] = [headers_out] + deduplicated_rows
        rows_out_count = len(deduplicated_rows)
        
        out = io.StringIO()
        writer = csv.writer(out, delimiter=delimiter, lineterminator="\n")
        writer.writerows(out_rows)
        
        finished = utc_now_iso()
        report = DataCleanReport(
            rows_in=len(data_rows),
            rows_out=rows_out_count,
            columns_in=len(raw_headers),
            columns_out=len(headers_out),
            header_map=header_map,
            fixes=fixes,
            started_at=started,
            finished_at=finished,
            duplicates_removed=fixes.get("duplicates_removed", 0),
            irrelevant_rows_removed=fixes.get("irrelevant_rows_removed", 0),
        )
        return out.getvalue(), report

    def clean_csv_file(self, input_path: str, output_path: str | None = None) -> DataCleanReport:
        with open(input_path, "r", encoding="utf-8-sig", errors="replace") as f:
            csv_text = f.read()
        cleaned, report = self.clean_csv_text(csv_text)
        if output_path is None:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}.cleaned{ext or '.csv'}"
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            f.write(cleaned)
        return report

    def detect_file_type(self, filename: str, content: bytes | None = None) -> str:
        """Detect file type from filename extension or content"""
        ext = os.path.splitext(filename.lower())[1] if filename else ""
        
        if ext in [".csv"]:
            return "csv"
        elif ext in [".xlsx", ".xls"]:
            return "excel"
        elif ext in [".json"]:
            return "json"
        elif ext in [".tsv"]:
            return "tsv"
        elif content:
            # Try to detect from content
            try:
                json.loads(content.decode("utf-8")[:100])
                return "json"
            except:
                pass
        return "csv"  # Default to CSV

    def detect_crm_type(self, headers: list[str]) -> str | None:
        """Detect CRM type from column headers"""
        headers_lower = [h.lower().strip() for h in headers]
        
        # Salesforce indicators
        sf_indicators = ["firstname", "lastname", "mailingstreet", "mailingcity", "createddate"]
        if any(ind in headers_lower for ind in sf_indicators):
            return "salesforce"
        
        # HubSpot indicators
        hs_indicators = ["firstname", "lastname", "lifecyclestage", "createdate", "hs_lastmodifieddate"]
        if any(ind in headers_lower for ind in hs_indicators):
            return "hubspot"
        
        # Pipedrive indicators
        pd_indicators = ["org_name", "owner_name", "add_time", "update_time"]
        if any(ind in headers_lower for ind in pd_indicators):
            return "pipedrive"
        
        return None

    def apply_crm_mappings(self, headers: list[str], crm_type: str) -> tuple[list[str], dict[str, str]]:
        """Apply CRM-specific field mappings to headers"""
        if crm_type not in self.CRM_FIELD_MAPPINGS:
            return headers, {}
        
        mappings = self.CRM_FIELD_MAPPINGS[crm_type]
        mapped_headers = []
        field_mappings = {}
        
        for h in headers:
            h_lower = h.lower().strip()
            if h_lower in mappings:
                mapped = mappings[h_lower]
                mapped_headers.append(mapped)
                field_mappings[h] = mapped
            else:
                mapped_headers.append(slugify_header(h))
                field_mappings[h] = slugify_header(h)
        
        return mapped_headers, field_mappings

    def parse_excel(self, file_content: bytes, sheet_name: str | None = None) -> tuple[list[list[str]], str]:
        """Parse Excel file (XLSX/XLS)"""
        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            raise ServiceError(
                "Excel support requires 'openpyxl'. Install with: pip install openpyxl"
            )
        
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = workbook.active if sheet_name is None else workbook[sheet_name]
        
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        
        return rows, "excel"
    
    def parse_json(self, file_content: bytes | str) -> tuple[list[list[str]], str]:
        """Parse JSON file and convert to tabular format"""
        if isinstance(file_content, bytes):
            content_str = file_content.decode("utf-8")
        else:
            content_str = file_content
        
        try:
            data = json.loads(content_str)
        except json.JSONDecodeError as e:
            raise ServiceError(f"Invalid JSON: {e}")
        
        if not isinstance(data, list):
            # If it's a single object or dict, wrap it
            if isinstance(data, dict):
                data = [data]
            else:
                raise ServiceError("JSON must be an array of objects or a single object")
        
        if not data:
            raise ServiceError("JSON appears to be empty")
        
        # Extract all unique keys as headers
        headers = []
        seen_keys = set()
        for item in data:
            if isinstance(item, dict):
                for key in item.keys():
                    if key not in seen_keys:
                        headers.append(str(key))
                        seen_keys.add(key)
        
        # Build rows
        rows = [headers]
        for item in data:
            if isinstance(item, dict):
                row = [str(item.get(h, "")) for h in headers]
                rows.append(row)
        
        return rows, "json"
    
    def parse_tsv(self, file_content: bytes | str) -> tuple[list[list[str]], str]:
        """Parse TSV (Tab-Separated Values) file"""
        if isinstance(file_content, bytes):
            content_str = file_content.decode("utf-8")
        else:
            content_str = file_content
        
        return self._parse_delimited(content_str, delimiter="\t"), "tsv"
    
    def _parse_delimited(self, content: str, delimiter: str = ",") -> list[list[str]]:
        """Parse delimited text (CSV/TSV)"""
        inp = io.StringIO(content)
        reader = csv.reader(inp, delimiter=delimiter)
        return list(reader)

    def clean_file(
        self,
        file_content: bytes,
        filename: str,
        *,
        file_type: str | None = None,
        delimiter: str = ",",
        normalize_headers: bool = True,
        drop_empty_rows: bool = True,
        apply_crm_mappings: bool = True,
        sheet_name: str | None = None,
    ) -> tuple[str, DataCleanReport]:
        """
        Clean a file of any supported type.
        Returns cleaned CSV text and report.
        """
        started = utc_now_iso()
        
        # Detect file type
        detected_type = file_type or self.detect_file_type(filename, file_content)
        
        # Parse file based on type
        if detected_type == "excel":
            rows, _ = self.parse_excel(file_content, sheet_name)
        elif detected_type == "json":
            rows, _ = self.parse_json(file_content)
        elif detected_type == "tsv":
            rows, _ = self.parse_tsv(file_content)
        else:  # CSV or default
            content_str = file_content.decode("utf-8") if isinstance(file_content, bytes) else file_content
            rows = self._parse_delimited(content_str, delimiter=delimiter)
        
        if not rows:
            raise ServiceError(f"{detected_type.upper()} file appears to be empty.")
        
        raw_headers = rows[0]
        data_rows = rows[1:]
        
        # Detect CRM type
        crm_type = self.detect_crm_type(raw_headers) if apply_crm_mappings else None
        
        # Apply CRM mappings if detected
        if crm_type and apply_crm_mappings:
            headers_out, field_mappings = self.apply_crm_mappings(raw_headers, crm_type)
        else:
            headers_out = [slugify_header(h) if normalize_headers else h.strip() for h in raw_headers]
            field_mappings = {h: headers_out[i] for i, h in enumerate(raw_headers)}
        
        # Use existing cleaning logic
        fixes: dict[str, int] = {
            "trimmed_cells": 0,
            "normalized_headers": 0,
            "normalized_dates": 0,
            "normalized_numbers": 0,
            "empties_to_blank": 0,
            "dropped_empty_rows": 0,
            "duplicates_removed": 0,
            "irrelevant_rows_removed": 0,
        }
        
        header_map: dict[str, str] = {h: headers_out[i] for i, h in enumerate(raw_headers)}
        for h in raw_headers:
            h2 = header_map[h]
            if h2 != h:
                fixes["normalized_headers"] += 1
        
        def norm_cell(val: str) -> str:
            if val is None:
                fixes["empties_to_blank"] += 1
                return ""
            v = str(val)
            v2 = v.strip()
            if v2 != v:
                fixes["trimmed_cells"] += 1
            v = v2
            if not v:
                return ""
            
            # numbers like "1,234.50" -> "1234.50"
            if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", v):
                fixes["normalized_numbers"] += 1
                return v.replace(",", "")
            
            # dates
            for pat in self._date_patterns:
                if pat.match(v):
                    iso = self._try_parse_date_to_iso(v)
                    if iso and iso != v:
                        fixes["normalized_dates"] += 1
                        return iso
                    return v
            return v
        
        def reconcile_row_length(row: list[str], target_cols: int) -> list[str]:
            if len(row) == target_cols:
                return row
            if len(row) < target_cols:
                return (row + [""] * target_cols)[:target_cols]
            
            rr = list(row)
            numeric_left = re.compile(r"^-?\d{1,3}$")
            numeric_mid = re.compile(r"^\d{3}$")
            numeric_right = re.compile(r"^\d{3}(\.\d+)?$")
            
            i = 0
            while len(rr) > target_cols and i < len(rr) - 1:
                a, b = rr[i].strip(), rr[i + 1].strip()
                if numeric_left.match(a) and (numeric_mid.match(b) or numeric_right.match(b)):
                    rr[i : i + 2] = [f"{a},{b}"]
                    fixes["normalized_numbers"] += 1
                    continue
                i += 1
            
            if len(rr) > target_cols:
                head = rr[: target_cols - 1]
                tail = rr[target_cols - 1 :]
                rr = head + [delimiter.join(tail)]
            return rr[:target_cols]
        
        # Process and clean rows
        cleaned_rows: list[list[str]] = []
        for r in data_rows:
            rr = reconcile_row_length(list(r), len(raw_headers))
            rr2 = [norm_cell(c) for c in rr]
            
            # Drop completely empty rows
            if drop_empty_rows and all(c == "" for c in rr2):
                fixes["dropped_empty_rows"] += 1
                continue
            
            # Filter irrelevant rows (rows with too many empty cells or test data)
            if self._is_irrelevant_row(rr2, len(headers_out)):
                fixes["irrelevant_rows_removed"] += 1
                continue
            
            cleaned_rows.append(rr2)
        
        # Remove duplicates
        seen_rows: set[tuple[str, ...]] = set()
        deduplicated_rows: list[list[str]] = []
        for row in cleaned_rows:
            # Create a tuple for hashing (normalize whitespace)
            row_tuple = tuple(c.strip().lower() if c else "" for c in row)
            if row_tuple not in seen_rows:
                seen_rows.add(row_tuple)
                deduplicated_rows.append(row)
            else:
                fixes["duplicates_removed"] += 1
        
        out_rows: list[list[str]] = [headers_out] + deduplicated_rows
        rows_out_count = len(deduplicated_rows)
        
        out = io.StringIO()
        writer = csv.writer(out, delimiter=delimiter, lineterminator="\n")
        writer.writerows(out_rows)
        
        finished = utc_now_iso()
        report = DataCleanReport(
            rows_in=len(data_rows),
            rows_out=rows_out_count,
            columns_in=len(raw_headers),
            columns_out=len(headers_out),
            header_map=header_map,
            fixes=fixes,
            started_at=started,
            finished_at=finished,
            file_type=detected_type,
            crm_detected=crm_type,
            field_mappings=field_mappings,
            duplicates_removed=fixes.get("duplicates_removed", 0),
            irrelevant_rows_removed=fixes.get("irrelevant_rows_removed", 0),
        )
        return out.getvalue(), report
    
    def _is_irrelevant_row(self, row: list[str], num_columns: int) -> bool:
        """
        Determine if a row is irrelevant and should be removed.
        Criteria:
        - More than 80% empty cells
        - Contains common test data indicators
        - Only contains whitespace or special characters
        """
        if not row:
            return True
        
        # Count non-empty cells
        non_empty_count = sum(1 for cell in row if cell and cell.strip())
        empty_percentage = (num_columns - non_empty_count) / max(num_columns, 1)
        
        # Remove rows with more than 80% empty cells
        if empty_percentage > 0.8:
            return True
        
        # Check for test data indicators
        test_indicators = [
            "test", "example", "sample", "dummy", "placeholder",
            "lorem ipsum", "xxx", "aaa", "123", "test@test.com"
        ]
        row_text = " ".join(cell.lower() for cell in row if cell)
        if any(indicator in row_text for indicator in test_indicators):
            # Only remove if it's clearly test data (not just containing the word)
            if any(row_text.startswith(ind) or row_text.endswith(ind) for ind in test_indicators):
                return True
        
        # Remove rows that are only whitespace/special characters
        if non_empty_count == 0:
            return True
        
        return False

    def _try_parse_date_to_iso(self, value: str) -> str | None:
        value = value.strip()
        # YYYY-MM-DD
        try:
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
                d = _dt.date.fromisoformat(value)
                return d.isoformat()
        except ValueError:
            pass
        # M/D/YYYY or M-D-YYYY
        m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", value)
        if m:
            mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if yy < 100:
                yy += 2000
            try:
                return _dt.date(yy, mm, dd).isoformat()
            except ValueError:
                return None
        # Salesforce datetime format: 2025-12-29T10:30:00.000Z
        try:
            if "T" in value:
                dt_str = value.split("T")[0]
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", dt_str):
                    d = _dt.date.fromisoformat(dt_str)
                    return d.isoformat()
        except ValueError:
            pass
        return None
    
    def clean_file_streaming(
        self,
        file_content: bytes,
        filename: str,
        *,
        file_type: str | None = None,
        delimiter: str = ",",
        normalize_headers: bool = True,
        drop_empty_rows: bool = True,
        apply_crm_mappings: bool = True,
        sheet_name: str | None = None,
        chunk_size: int = 10000,
        export_formats: list[str] | None = None,
    ) -> tuple[dict[str, t.Any], DataCleanReport]:
        """
        Clean large files using streaming/chunked processing to handle millions of rows.
        Returns a dict with multiple output formats and column-based files.
        """
        started = utc_now_iso()
        
        # Detect file type
        detected_type = file_type or self.detect_file_type(filename, file_content)
        
        # Parse file based on type
        if detected_type == "excel":
            rows, _ = self.parse_excel(file_content, sheet_name)
        elif detected_type == "json":
            rows, _ = self.parse_json(file_content)
        elif detected_type == "tsv":
            rows, _ = self.parse_tsv(file_content)
        else:  # CSV or default
            content_str = file_content.decode("utf-8") if isinstance(file_content, bytes) else file_content
            rows = self._parse_delimited(content_str, delimiter=delimiter)
        
        if not rows:
            raise ServiceError(f"{detected_type.upper()} file appears to be empty.")
        
        raw_headers = rows[0]
        data_rows = rows[1:]
        
        # For very large files, process in chunks
        if len(data_rows) > chunk_size:
            return self._process_large_file_chunked(
                rows, raw_headers, detected_type, delimiter, normalize_headers,
                drop_empty_rows, apply_crm_mappings, started, chunk_size, export_formats
            )
        
        # Use regular processing for smaller files
        cleaned_csv, report = self.clean_file(
            file_content, filename,
            file_type=file_type,
            delimiter=delimiter,
            normalize_headers=normalize_headers,
            drop_empty_rows=drop_empty_rows,
            apply_crm_mappings=apply_crm_mappings,
            sheet_name=sheet_name,
        )
        
        # Parse cleaned CSV to get rows for multiple outputs
        reader = csv.reader(io.StringIO(cleaned_csv))
        rows_list = list(reader)
        headers_out = rows_list[0] if rows_list else raw_headers
        cleaned_rows = rows_list[1:] if len(rows_list) > 1 else []
        
        # Generate multiple output formats (only requested ones)
        export_formats = export_formats or ['csv', 'json', 'excel', 'columns']
        outputs = self._generate_multiple_outputs(
            cleaned_csv, raw_headers, detected_type, report, cleaned_rows, headers_out, export_formats
        )
        
        return outputs, report
    
    def _process_large_file_chunked(
        self,
        rows: list[list[str]],
        raw_headers: list[str],
        detected_type: str,
        delimiter: str,
        normalize_headers: bool,
        drop_empty_rows: bool,
        apply_crm_mappings: bool,
        started: str,
        chunk_size: int = 10000,
        export_formats: list[str] | None = None,
    ) -> tuple[dict[str, t.Any], DataCleanReport]:
        """Process very large files in chunks to avoid memory issues"""
        fixes: dict[str, int] = {
            "trimmed_cells": 0,
            "normalized_headers": 0,
            "normalized_dates": 0,
            "normalized_numbers": 0,
            "empties_to_blank": 0,
            "dropped_empty_rows": 0,
            "duplicates_removed": 0,
            "irrelevant_rows_removed": 0,
        }
        
        data_rows = rows[1:]
        
        # Detect CRM type
        crm_type = self.detect_crm_type(raw_headers) if apply_crm_mappings else None
        
        # Apply CRM mappings if detected
        if crm_type and apply_crm_mappings:
            headers_out, field_mappings = self.apply_crm_mappings(raw_headers, crm_type)
        else:
            headers_out = [slugify_header(h) if normalize_headers else h.strip() for h in raw_headers]
            field_mappings = {h: headers_out[i] for i, h in enumerate(raw_headers)}
        
        header_map: dict[str, str] = {h: headers_out[i] for i, h in enumerate(raw_headers)}
        
        # Process in chunks
        all_cleaned_rows: list[list[str]] = []
        seen_rows: set[tuple[str, ...]] = set()
        
        total_chunks = (len(data_rows) + chunk_size - 1) // chunk_size
        
        for chunk_idx in range(total_chunks):
            start_idx = chunk_idx * chunk_size
            end_idx = min(start_idx + chunk_size, len(data_rows))
            chunk = data_rows[start_idx:end_idx]
            
            # Process chunk
            for r in chunk:
                rr = self._reconcile_row_length(list(r), len(raw_headers), delimiter, fixes)
                rr2 = [self._norm_cell(c, fixes) for c in rr]
                
                if drop_empty_rows and all(c == "" for c in rr2):
                    fixes["dropped_empty_rows"] += 1
                    continue
                
                if self._is_irrelevant_row(rr2, len(headers_out)):
                    fixes["irrelevant_rows_removed"] += 1
                    continue
                
                # Check for duplicates
                row_tuple = tuple(c.strip().lower() if c else "" for c in rr2)
                if row_tuple not in seen_rows:
                    seen_rows.add(row_tuple)
                    all_cleaned_rows.append(rr2)
                else:
                    fixes["duplicates_removed"] += 1
        
        # Generate outputs
        cleaned_csv = self._rows_to_csv([headers_out] + all_cleaned_rows, delimiter)
        report = DataCleanReport(
            rows_in=len(data_rows),
            rows_out=len(all_cleaned_rows),
            columns_in=len(raw_headers),
            columns_out=len(headers_out),
            header_map=header_map,
            fixes=fixes,
            started_at=started,
            finished_at=utc_now_iso(),
            file_type=detected_type,
            crm_detected=crm_type,
            field_mappings=field_mappings,
            duplicates_removed=fixes.get("duplicates_removed", 0),
            irrelevant_rows_removed=fixes.get("irrelevant_rows_removed", 0),
        )
        
        # For very large files, optimize export generation
        export_formats = export_formats or ['csv', 'json', 'excel', 'columns']
        num_rows = len(all_cleaned_rows)
        
        # For files with 50k+ rows, limit exports to prevent memory issues
        if num_rows > 50000:
            # Only generate CSV and skip memory-intensive formats unless explicitly requested
            safe_formats = ['csv']
            if 'json' in export_formats and num_rows <= 100000:
                safe_formats.append('json')
            # Excel and column files are too memory-intensive for very large files
            if 'excel' in export_formats and num_rows <= 50000:
                safe_formats.append('excel')
            # Column files only for smaller large files
            if 'columns' in export_formats and num_rows <= 25000:
                safe_formats.append('columns')
            
            export_formats = safe_formats
        
        outputs = self._generate_multiple_outputs_optimized(
            cleaned_csv, raw_headers, detected_type, report, all_cleaned_rows, headers_out, export_formats, num_rows
        )
        
        return outputs, report
    
    def _generate_multiple_outputs_optimized(
        self,
        cleaned_csv: str,
        raw_headers: list[str],
        original_file_type: str,
        report: DataCleanReport,
        cleaned_rows: list[list[str]] | None = None,
        headers_out: list[str] | None = None,
        export_formats: list[str] | None = None,
        num_rows: int = 0,
    ) -> dict[str, t.Any]:
        """Generate multiple output formats with optimizations for large files"""
        export_formats = export_formats or ['csv', 'json', 'excel', 'columns']
        outputs: dict[str, t.Any] = {}
        
        # Parse cleaned CSV to get rows if not provided
        if cleaned_rows is None or headers_out is None:
            reader = csv.reader(io.StringIO(cleaned_csv))
            rows_list = list(reader)
            if rows_list:
                headers_out = rows_list[0]
                cleaned_rows = rows_list[1:]
            else:
                return outputs
        
        # Generate CSV output (always include as base - most memory efficient)
        if 'csv' in export_formats:
            outputs["master_cleanse_csv"] = cleaned_csv
        
        # Generate JSON output (memory-intensive for large files)
        if 'json' in export_formats:
            if num_rows <= 100000:
                json_output = self._rows_to_json(cleaned_rows, headers_out)
                outputs["master_cleanse_json"] = json_output
            else:
                # For very large files, skip JSON or provide a note
                outputs["master_cleanse_json"] = None
                outputs["_json_skipped"] = f"JSON export skipped for files with {num_rows:,} rows to prevent memory issues"
        
        # Generate Excel output (memory-intensive for large files)
        if 'excel' in export_formats:
            if num_rows <= 50000:
                try:
                    excel_output = self._rows_to_excel(cleaned_rows, headers_out)
                    outputs["master_cleanse_excel"] = excel_output
                except Exception as e:
                    outputs["master_cleanse_excel"] = None
                    outputs["_excel_error"] = str(e)
            else:
                outputs["master_cleanse_excel"] = None
                outputs["_excel_skipped"] = f"Excel export skipped for files with {num_rows:,} rows to prevent memory issues"
        
        # Generate column-based files (very memory-intensive for large files)
        if 'columns' in export_formats:
            if num_rows <= 25000:
                column_files = self._generate_column_based_files(cleaned_rows, headers_out, original_file_type, export_formats)
                outputs["column_files"] = column_files
            else:
                # For very large files, only generate CSV column files
                column_files = self._generate_column_based_files(cleaned_rows, headers_out, original_file_type, ['csv'])
                outputs["column_files"] = column_files
                outputs["_columns_note"] = f"Column files limited to CSV format for files with {num_rows:,} rows"
        
        return outputs
    
    def _generate_multiple_outputs(
        self,
        cleaned_csv: str,
        raw_headers: list[str],
        original_file_type: str,
        report: DataCleanReport,
        cleaned_rows: list[list[str]] | None = None,
        headers_out: list[str] | None = None,
        export_formats: list[str] | None = None,
    ) -> dict[str, t.Any]:
        """Generate multiple output formats: CSV, JSON, Excel, and column-based files (only requested formats)"""
        export_formats = export_formats or ['csv', 'json', 'excel', 'columns']
        outputs: dict[str, t.Any] = {}
        
        # Parse cleaned CSV to get rows if not provided
        if cleaned_rows is None or headers_out is None:
            reader = csv.reader(io.StringIO(cleaned_csv))
            rows_list = list(reader)
            if rows_list:
                headers_out = rows_list[0]
                cleaned_rows = rows_list[1:]
            else:
                return outputs
        
        # Generate CSV output (always include as base)
        if 'csv' in export_formats:
            outputs["master_cleanse_csv"] = cleaned_csv
        
        # Generate JSON output
        if 'json' in export_formats:
            json_output = self._rows_to_json(cleaned_rows, headers_out)
            outputs["master_cleanse_json"] = json_output
        
        # Generate Excel output
        if 'excel' in export_formats:
            try:
                excel_output = self._rows_to_excel(cleaned_rows, headers_out)
                outputs["master_cleanse_excel"] = excel_output
            except Exception:
                # Excel generation failed, skip it
                pass
        
        # Always generate column-based files (one per column) - this is a core feature
        # Column files help ensure no accidental deletions during data verification
        column_files = self._generate_column_based_files(cleaned_rows, headers_out, original_file_type, export_formats)
        outputs["column_files"] = column_files
        
        return outputs
    
    def _rows_to_json(self, rows: list[list[str]], headers: list[str]) -> str:
        """Convert rows to JSON format - optimized for large files"""
        # For very large files, use streaming JSON generation
        if len(rows) > 50000:
            return self._rows_to_json_streaming(rows, headers)
        
        data = []
        for row in rows:
            obj = {}
            for i, header in enumerate(headers):
                obj[header] = row[i] if i < len(row) else ""
            data.append(obj)
        return json.dumps(data, indent=2, ensure_ascii=False)
    
    def _rows_to_json_streaming(self, rows: list[list[str]], headers: list[str]) -> str:
        """Stream JSON generation for very large files to reduce memory usage"""
        output = io.StringIO()
        output.write('[\n')
        
        for idx, row in enumerate(rows):
            if idx > 0:
                output.write(',\n')
            obj = {}
            for i, header in enumerate(headers):
                obj[header] = row[i] if i < len(row) else ""
            json.dump(obj, output, ensure_ascii=False)
        
        output.write('\n]')
        return output.getvalue()
    
    def _rows_to_excel(self, rows: list[list[str]], headers: list[str]) -> bytes:
        """Convert rows to Excel format - optimized for large files"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ServiceError("Excel export requires 'openpyxl'. Install with: pip install openpyxl")
        
        # Use write-only mode for large files to reduce memory usage
        output = io.BytesIO()
        
        if len(rows) > 10000:
            # Use write-only workbook for large files
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Cleaned Data")
            
            # Write headers
            ws.append(headers)
            
            # Write rows in chunks
            for row in rows:
                ws.append(row)
        else:
            # Use regular workbook for smaller files
            wb = Workbook()
            ws = wb.active
            ws.title = "Cleaned Data"
            
            # Write headers
            ws.append(headers)
            
            # Write rows
            for row in rows:
                ws.append(row)
        
        # Save to bytes
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _generate_column_based_files(
        self,
        rows: list[list[str]],
        headers: list[str],
        original_file_type: str,
        export_formats: list[str] | None = None,
    ) -> dict[str, dict[str, str | bytes]]:
        """Generate one file per column to ensure no accidental deletions - optimized for large files"""
        export_formats = export_formats or ['csv', 'json', 'excel']
        column_files: dict[str, dict[str, str | bytes]] = {}
        
        num_rows = len(rows)
        
        # For very large files, only generate CSV column files to save memory
        if num_rows > 50000:
            export_formats = ['csv']
        
        for col_idx, header in enumerate(headers):
            # Extract column data efficiently
            column_data = [row[col_idx] if col_idx < len(row) else "" for row in rows]
            column_files[header] = {}
            
            # Create CSV file for this column (always available)
            if 'csv' in export_formats:
                csv_rows = [[header]] + [[val] for val in column_data]
                csv_content = self._rows_to_csv(csv_rows, ",")
                column_files[header]["csv"] = csv_content
            
            # Create JSON file for this column (only for smaller files)
            if 'json' in export_formats and num_rows <= 25000:
                json_data = [{header: val} for val in column_data]
                json_content = json.dumps(json_data, indent=2, ensure_ascii=False)
                column_files[header]["json"] = json_content
            
            # Create Excel file for this column (only for smaller files)
            if 'excel' in export_formats and num_rows <= 10000:
                try:
                    csv_rows = [[header]] + [[val] for val in column_data]
                    excel_content = self._rows_to_excel(csv_rows[1:], [header])
                    column_files[header]["excel"] = excel_content
                except:
                    column_files[header]["excel"] = None
        
        return column_files
    
    def _rows_to_csv(self, rows: list[list[str]], delimiter: str) -> str:
        """Convert rows to CSV string"""
        out = io.StringIO()
        writer = csv.writer(out, delimiter=delimiter, lineterminator="\n")
        writer.writerows(rows)
        return out.getvalue()
    
    def _norm_cell(self, val: str, fixes: dict[str, int]) -> str:
        """Normalize a cell value (helper for chunked processing)"""
        if val is None:
            fixes["empties_to_blank"] += 1
            return ""
        v = str(val)
        v2 = v.strip()
        if v2 != v:
            fixes["trimmed_cells"] += 1
        v = v2
        if not v:
            return ""
        
        # numbers like "1,234.50" -> "1234.50"
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", v):
            fixes["normalized_numbers"] += 1
            return v.replace(",", "")
        
        # dates
        for pat in self._date_patterns:
            if pat.match(v):
                iso = self._try_parse_date_to_iso(v)
                if iso and iso != v:
                    fixes["normalized_dates"] += 1
                    return iso
                return v
        return v
    
    def _reconcile_row_length(
        self,
        row: list[str],
        target_cols: int,
        delimiter: str,
        fixes: dict[str, int],
    ) -> list[str]:
        """Reconcile row length (helper for chunked processing)"""
        if len(row) == target_cols:
            return row
        if len(row) < target_cols:
            return (row + [""] * target_cols)[:target_cols]
        
        rr = list(row)
        numeric_left = re.compile(r"^-?\d{1,3}$")
        numeric_mid = re.compile(r"^\d{3}$")
        numeric_right = re.compile(r"^\d{3}(\.\d+)?$")
        
        i = 0
        while len(rr) > target_cols and i < len(rr) - 1:
            a, b = rr[i].strip(), rr[i + 1].strip()
            if numeric_left.match(a) and (numeric_mid.match(b) or numeric_right.match(b)):
                rr[i : i + 2] = [f"{a},{b}"]
                fixes["normalized_numbers"] += 1
                continue
            i += 1
        
        if len(rr) > target_cols:
            head = rr[: target_cols - 1]
            tail = rr[target_cols - 1 :]
            rr = head + [delimiter.join(tail)]
        return rr[:target_cols]


